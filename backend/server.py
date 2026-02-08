from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, UploadFile, File, Response
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import openpyxl
from io import BytesIO
import base64
import hashlib
import secrets
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import ftplib
import asyncio
from contextlib import asynccontextmanager

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'magazyn_db')]

# Create the main app
app = FastAPI(title="Magazyn ITS Kielce API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== HELPER FUNCTIONS ====================

def hash_password(password: str) -> str:
    """Hash password with SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

def generate_token() -> str:
    """Generate a secure random token"""
    return secrets.token_urlsafe(32)

# ==================== MODELS ====================

class LoginRequest(BaseModel):
    email: str
    password: str

class CreateUserRequest(BaseModel):
    email: str
    password: str
    name: str
    role: str = "pracownik"

class ChangePasswordRequest(BaseModel):
    current_password: Optional[str] = None  # Required for self-change
    new_password: str

class Device(BaseModel):
    device_id: str = Field(default_factory=lambda: f"dev_{uuid.uuid4().hex[:12]}")
    nazwa: str
    numer_seryjny: str
    kod_kreskowy: Optional[str] = None
    kod_qr: Optional[str] = None
    przypisany_do: Optional[str] = None
    status: str = "dostepny"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DeviceInstallation(BaseModel):
    installation_id: str = Field(default_factory=lambda: f"inst_{uuid.uuid4().hex[:12]}")
    device_id: str
    user_id: str
    nazwa_urzadzenia: str
    data_instalacji: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    adres: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    rodzaj_zlecenia: str

class Message(BaseModel):
    message_id: str = Field(default_factory=lambda: f"msg_{uuid.uuid4().hex[:12]}")
    sender_id: str
    sender_name: str
    content: Optional[str] = None
    attachment: Optional[str] = None
    attachment_type: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Task(BaseModel):
    task_id: str = Field(default_factory=lambda: f"task_{uuid.uuid4().hex[:12]}")
    title: str
    description: Optional[str] = None
    assigned_to: str
    assigned_by: str
    due_date: datetime
    status: str = "oczekujace"
    priority: str = "normalne"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BackupSettings(BaseModel):
    # Email settings
    smtp_host: Optional[str] = None
    smtp_port: Optional[int] = 587
    smtp_user: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_use_tls: bool = True
    email_recipient: Optional[str] = None
    email_enabled: bool = False
    
    # FTP settings
    ftp_host: Optional[str] = None
    ftp_port: Optional[int] = 21
    ftp_user: Optional[str] = None
    ftp_password: Optional[str] = None
    ftp_path: Optional[str] = "/backups/"
    ftp_enabled: bool = False
    
    # Schedule settings
    schedule_enabled: bool = False
    schedule_time: str = "02:00"  # HH:MM format
    
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BackupLog(BaseModel):
    backup_id: str = Field(default_factory=lambda: f"backup_{uuid.uuid4().hex[:12]}")
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    size_bytes: int
    status: str  # "success", "failed"
    sent_email: bool = False
    sent_ftp: bool = False
    error_message: Optional[str] = None

class ActivityLog(BaseModel):
    """Model for tracking all user activities and device history"""
    log_id: str = Field(default_factory=lambda: f"log_{uuid.uuid4().hex[:12]}")
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
    # Who performed the action
    user_id: str
    user_name: str
    user_role: str
    
    # What action was performed
    action_type: str  # login, logout, device_scan, device_install, device_assign, device_return, task_create, task_complete, etc.
    action_description: str
    
    # Related entities
    device_serial: Optional[str] = None
    device_name: Optional[str] = None
    device_id: Optional[str] = None
    task_id: Optional[str] = None
    target_user_id: Optional[str] = None
    target_user_name: Optional[str] = None
    
    # Additional details
    details: Optional[dict] = None
    ip_address: Optional[str] = None

# ==================== ACTIVITY LOGGING HELPER ====================

async def log_activity(
    user_id: str,
    user_name: str,
    user_role: str,
    action_type: str,
    action_description: str,
    device_serial: str = None,
    device_name: str = None,
    device_id: str = None,
    task_id: str = None,
    target_user_id: str = None,
    target_user_name: str = None,
    details: dict = None,
    ip_address: str = None
):
    """Log user activity to the database"""
    log_entry = {
        "log_id": f"log_{uuid.uuid4().hex[:12]}",
        "timestamp": datetime.now(timezone.utc),
        "user_id": user_id,
        "user_name": user_name,
        "user_role": user_role,
        "action_type": action_type,
        "action_description": action_description,
        "device_serial": device_serial,
        "device_name": device_name,
        "device_id": device_id,
        "task_id": task_id,
        "target_user_id": target_user_id,
        "target_user_name": target_user_name,
        "details": details,
        "ip_address": ip_address
    }
    await db.activity_logs.insert_one(log_entry)
    return log_entry

# ==================== AUTH HELPERS ====================

async def get_session_token(request: Request) -> Optional[str]:
    """Extract session token from cookies or Authorization header"""
    session_token = request.cookies.get("session_token")
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header[7:]
    return session_token

async def get_current_user(request: Request) -> Optional[dict]:
    """Get current user from session"""
    session_token = await get_session_token(request)
    if not session_token:
        return None
    
    session = await db.user_sessions.find_one(
        {"session_token": session_token},
        {"_id": 0}
    )
    if not session:
        return None
    
    expires_at = session["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at <= datetime.now(timezone.utc):
        return None
    
    user_doc = await db.users.find_one(
        {"user_id": session["user_id"]},
        {"_id": 0, "password_hash": 0}
    )
    return user_doc

async def require_user(request: Request) -> dict:
    """Require authenticated user"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Nie zalogowany")
    return user

async def require_admin(request: Request) -> dict:
    """Require admin user"""
    user = await require_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Brak uprawnień administratora")
    return user

# ==================== STARTUP - CREATE ADMIN ====================

@app.on_event("startup")
async def create_default_admin():
    """Create default admin account if not exists"""
    admin_email = "kamil@magazyn.its.kielce.pl"
    existing = await db.users.find_one({"email": admin_email})
    
    if not existing:
        admin_user = {
            "user_id": f"user_{uuid.uuid4().hex[:12]}",
            "email": admin_email,
            "name": "Kamil",
            "password_hash": hash_password("kamil678@"),
            "role": "admin",
            "created_at": datetime.now(timezone.utc)
        }
        await db.users.insert_one(admin_user)
        logger.info(f"Created default admin account: {admin_email}")

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/login")
async def login(data: LoginRequest, request: Request, response: Response):
    """Login with email and password"""
    user = await db.users.find_one(
        {"email": data.email.lower()},
        {"_id": 0}
    )
    
    if not user:
        raise HTTPException(status_code=401, detail="Nieprawidłowy email lub hasło")
    
    password_hash = hash_password(data.password)
    if user.get("password_hash") != password_hash:
        raise HTTPException(status_code=401, detail="Nieprawidłowy email lub hasło")
    
    # Delete old sessions
    await db.user_sessions.delete_many({"user_id": user["user_id"]})
    
    # Create new session
    session_token = generate_token()
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    # Get IP address and User-Agent
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    client_ip = forwarded_for.split(",")[0].strip() if forwarded_for else (request.client.host if request.client else "nieznany")
    user_agent = request.headers.get("User-Agent", "nieznany")
    
    await db.user_sessions.insert_one({
        "user_id": user["user_id"],
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    })
    
    # Save last login info to user document
    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {
            "last_login_at": datetime.now(timezone.utc),
            "last_login_ip": client_ip,
            "last_login_device": user_agent
        }}
    )
    
    # Log login activity
    await log_activity(
        user_id=user["user_id"],
        user_name=user["name"],
        user_role=user.get("role", "pracownik"),
        action_type="login",
        action_description=f"Zalogowano do systemu",
        ip_address=client_ip,
        details={"device": user_agent}
    )
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 60 * 60
    )
    
    return {
        "user_id": user["user_id"],
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
        "session_token": session_token
    }

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(require_user)):
    """Get current user"""
    return user

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    """Logout user"""
    session_token = await get_session_token(request)
    if session_token:
        await db.user_sessions.delete_many({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Wylogowano pomyślnie"}

@api_router.post("/auth/change-password")
async def change_own_password(request: Request, user: dict = Depends(require_user)):
    """Change own password"""
    body = await request.json()
    current_password = body.get("current_password")
    new_password = body.get("new_password")
    
    if not current_password or not new_password:
        raise HTTPException(status_code=400, detail="Wymagane aktualne i nowe hasło")
    
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Hasło musi mieć minimum 6 znaków")
    
    # Verify current password
    user_doc = await db.users.find_one({"user_id": user["user_id"]})
    if user_doc.get("password_hash") != hash_password(current_password):
        raise HTTPException(status_code=401, detail="Nieprawidłowe aktualne hasło")
    
    # Update password
    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {"password_hash": hash_password(new_password)}}
    )
    
    return {"message": "Hasło zostało zmienione"}

# ==================== USER MANAGEMENT (ADMIN) ====================

@api_router.get("/users")
async def get_users(admin: dict = Depends(require_admin)):
    """Get all users (admin only)"""
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.post("/users")
async def create_user(data: CreateUserRequest, admin: dict = Depends(require_admin)):
    """Create new user (admin only)"""
    # Check if email already exists
    existing = await db.users.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email jest już zarejestrowany")
    
    if len(data.password) < 6:
        raise HTTPException(status_code=400, detail="Hasło musi mieć minimum 6 znaków")
    
    if data.role not in ["admin", "pracownik"]:
        raise HTTPException(status_code=400, detail="Nieprawidłowa rola")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    
    user_doc = {
        "user_id": user_id,
        "email": data.email.lower(),
        "name": data.name,
        "password_hash": hash_password(data.password),
        "role": data.role,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.users.insert_one(user_doc)
    
    return {
        "user_id": user_id,
        "email": data.email.lower(),
        "name": data.name,
        "role": data.role,
        "message": "Użytkownik został utworzony"
    }

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, request: Request, admin: dict = Depends(require_admin)):
    """Update user role (admin only)"""
    body = await request.json()
    new_role = body.get("role")
    
    if new_role not in ["admin", "pracownik"]:
        raise HTTPException(status_code=400, detail="Nieprawidłowa rola")
    
    result = await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": new_role}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono użytkownika")
    
    return {"message": "Rola zaktualizowana"}

@api_router.put("/users/{user_id}/password")
async def reset_user_password(user_id: str, request: Request, admin: dict = Depends(require_admin)):
    """Reset user password (admin only)"""
    body = await request.json()
    new_password = body.get("new_password")
    
    if not new_password or len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Hasło musi mieć minimum 6 znaków")
    
    result = await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"password_hash": hash_password(new_password)}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono użytkownika")
    
    # Invalidate user sessions
    await db.user_sessions.delete_many({"user_id": user_id})
    
    return {"message": "Hasło zostało zresetowane"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    """Delete user (admin only)"""
    if user_id == admin["user_id"]:
        raise HTTPException(status_code=400, detail="Nie możesz usunąć własnego konta")
    
    result = await db.users.delete_one({"user_id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono użytkownika")
    
    # Delete user sessions
    await db.user_sessions.delete_many({"user_id": user_id})
    
    return {"message": "Użytkownik został usunięty"}

@api_router.get("/workers")
async def get_workers(user: dict = Depends(require_user)):
    """Get all workers (pracownik role)"""
    workers = await db.users.find({"role": "pracownik"}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return workers

# ==================== DEVICE MANAGEMENT ====================

@api_router.post("/devices/import")
async def import_devices(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    """Import devices from XLSX file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Tylko pliki XLSX są obsługiwane")
    
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    
    devices_imported = 0
    duplicates = 0
    errors = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        
        try:
            numer_seryjny = str(row[1]) if len(row) > 1 and row[1] else ""
            
            if not numer_seryjny:
                errors.append(f"Wiersz {row_num}: Brak numeru seryjnego")
                continue
            
            # Check for duplicate in database
            existing = await db.devices.find_one({"numer_seryjny": numer_seryjny})
            if existing:
                duplicates += 1
                errors.append(f"Wiersz {row_num}: Numer seryjny {numer_seryjny} już istnieje w systemie")
                continue
            
            device = {
                "device_id": f"dev_{uuid.uuid4().hex[:12]}",
                "nazwa": str(row[0]) if row[0] else "",
                "numer_seryjny": numer_seryjny,
                "kod_kreskowy": str(row[2]) if len(row) > 2 and row[2] else None,
                "kod_qr": str(row[3]) if len(row) > 3 and row[3] else None,
                "przypisany_do": None,
                "status": "dostepny",
                "created_at": datetime.now(timezone.utc),
                "imported_at": datetime.now(timezone.utc),
                "imported_by": admin["user_id"]
            }
            
            await db.devices.insert_one(device)
            
            # Log device import
            await log_activity(
                user_id=admin["user_id"],
                user_name=admin["name"],
                user_role="admin",
                action_type="device_import",
                action_description=f"Zaimportowano urządzenie {device['nazwa']} ({numer_seryjny}) z pliku Excel",
                device_serial=numer_seryjny,
                device_name=device["nazwa"],
                device_id=device["device_id"],
                details={"source": "excel_import", "filename": file.filename}
            )
            
            devices_imported += 1
        except Exception as e:
            errors.append(f"Wiersz {row_num}: {str(e)}")
    
    return {
        "imported": devices_imported,
        "duplicates": duplicates,
        "errors": errors,
        "message": f"Zaimportowano {devices_imported} urządzeń" + (f", pominięto {duplicates} duplikatów" if duplicates > 0 else "")
    }

@api_router.get("/devices")
async def get_devices(
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    user: dict = Depends(require_user)
):
    """Get all devices - workers can only see their own devices"""
    is_admin = user.get("role") == "admin"
    
    if is_admin:
        # Admin can see all devices
        query = {}
        if status:
            query["status"] = status
        if assigned_to:
            query["przypisany_do"] = assigned_to
        
        devices = await db.devices.find(query, {"_id": 0}).to_list(1000)
    else:
        # Worker - get assigned devices + devices they installed
        worker_id = user["user_id"]
        
        if status == "zainstalowany":
            # For installed devices, find devices this worker installed
            installations = await db.installations.find(
                {"user_id": worker_id},
                {"_id": 0}
            ).to_list(1000)
            
            installed_device_ids = [inst["device_id"] for inst in installations]
            devices = await db.devices.find(
                {"device_id": {"$in": installed_device_ids}, "status": "zainstalowany"},
                {"_id": 0}
            ).to_list(1000)
        elif status:
            # Other statuses - only assigned devices
            devices = await db.devices.find(
                {"przypisany_do": worker_id, "status": status},
                {"_id": 0}
            ).to_list(1000)
        else:
            # All devices - assigned + installed by this worker
            assigned_devices = await db.devices.find(
                {"przypisany_do": worker_id},
                {"_id": 0}
            ).to_list(1000)
            
            # Get installations by this worker
            installations = await db.installations.find(
                {"user_id": worker_id},
                {"_id": 0}
            ).to_list(1000)
            installed_device_ids = [inst["device_id"] for inst in installations]
            
            installed_devices = await db.devices.find(
                {"device_id": {"$in": installed_device_ids}, "status": "zainstalowany"},
                {"_id": 0}
            ).to_list(1000)
            
            # Merge without duplicates
            device_ids = set()
            devices = []
            for d in assigned_devices + installed_devices:
                if d["device_id"] not in device_ids:
                    device_ids.add(d["device_id"])
                    devices.append(d)
    
    # For installed devices, add installation info
    installed_device_ids = [d["device_id"] for d in devices if d.get("status") == "zainstalowany"]
    if installed_device_ids:
        installations = await db.installations.find(
            {"device_id": {"$in": installed_device_ids}},
            {"_id": 0}
        ).sort("data_instalacji", -1).to_list(1000)
        
        # Create a map of device_id to latest installation
        installation_map = {}
        for inst in installations:
            device_id = inst.get("device_id")
            if device_id and device_id not in installation_map:
                installation_map[device_id] = inst
        
        # Add installation info to devices
        for device in devices:
            if device.get("status") == "zainstalowany" and device["device_id"] in installation_map:
                inst = installation_map[device["device_id"]]
                device["instalacja"] = {
                    "adres": inst.get("adres") or inst.get("adres_klienta"),
                    "data_instalacji": inst.get("data_instalacji"),
                    "rodzaj_zlecenia": inst.get("rodzaj_zlecenia"),
                    "instalator_id": inst.get("user_id")
                }
    
    return devices

@api_router.get("/devices/inventory/summary")
async def get_inventory_summary(admin: dict = Depends(require_admin)):
    """Get inventory summary for all users (admin only)"""
    # Get all users
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    
    inventory = []
    for usr in users:
        # Get devices assigned to this user (przypisane)
        assigned_devices = await db.devices.find(
            {"przypisany_do": usr["user_id"], "status": "przypisany"},
            {"_id": 0}
        ).to_list(1000)
        
        # Get installed devices
        installed_devices = await db.devices.find(
            {"przypisany_do": usr["user_id"], "status": "zainstalowany"},
            {"_id": 0}
        ).to_list(1000)
        
        # Get damaged devices by this user
        damaged_devices = await db.devices.find(
            {"damaged_by": usr["user_id"], "status": "uszkodzony"},
            {"_id": 0}
        ).to_list(1000)
        
        # Count by barcode for assigned devices
        barcode_counts = {}
        for device in assigned_devices:
            barcode = device.get("kod_kreskowy") or device.get("nazwa") or "brak_kodu"
            if barcode not in barcode_counts:
                barcode_counts[barcode] = {
                    "kod_kreskowy": barcode,
                    "nazwa": device.get("nazwa", ""),
                    "count": 0,
                    "devices": []
                }
            barcode_counts[barcode]["count"] += 1
            barcode_counts[barcode]["devices"].append(device)
        
        # Check for low stock (less than 4 of same barcode)
        low_stock_items = [
            item for item in barcode_counts.values() 
            if item["count"] < 4
        ]
        
        inventory.append({
            "user_id": usr["user_id"],
            "user_name": usr["name"],
            "user_email": usr["email"],
            "role": usr["role"],
            "total_devices": len(assigned_devices),
            "total_installed": len(installed_devices),
            "total_damaged": len(damaged_devices),
            "by_barcode": list(barcode_counts.values()),
            "low_stock": low_stock_items,
            "has_low_stock": len(low_stock_items) > 0
        })
    
    return inventory

@api_router.get("/devices/inventory/{user_id}")
async def get_user_inventory(user_id: str, admin: dict = Depends(require_admin)):
    """Get detailed inventory for a specific user (admin only)"""
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="Nie znaleziono użytkownika")
    
    devices = await db.devices.find(
        {"przypisany_do": user_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Also get devices damaged by this user
    damaged_devices = await db.devices.find(
        {"damaged_by": user_id, "status": "uszkodzony"},
        {"_id": 0}
    ).to_list(1000)
    
    # Separate by status
    available = [d for d in devices if d.get("status") == "przypisany"]
    installed = [d for d in devices if d.get("status") == "zainstalowany"]
    damaged = damaged_devices
    
    # Count by barcode for available devices
    barcode_counts = {}
    for device in available:
        barcode = device.get("kod_kreskowy") or device.get("nazwa") or "brak_kodu"
        if barcode not in barcode_counts:
            barcode_counts[barcode] = {
                "kod_kreskowy": barcode,
                "nazwa": device.get("nazwa", ""),
                "count": 0,
                "devices": []
            }
        barcode_counts[barcode]["count"] += 1
        barcode_counts[barcode]["devices"].append(device)
    
    # Check for low stock items
    low_stock = [item for item in barcode_counts.values() if item["count"] < 4]
    
    return {
        "user": target_user,
        "total_available": len(available),
        "total_installed": len(installed),
        "total_damaged": len(damaged),
        "available_devices": available,
        "installed_devices": installed,
        "damaged_devices": damaged,
        "by_barcode": list(barcode_counts.values()),
        "low_stock": low_stock,
        "has_low_stock": len(low_stock) > 0
    }

@api_router.get("/devices/{device_id}")
async def get_device(device_id: str, user: dict = Depends(require_user)):
    """Get single device"""
    device = await db.devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Nie znaleziono urządzenia")
    return device

@api_router.post("/devices/{device_id}/assign")
async def assign_device(device_id: str, request: Request, admin: dict = Depends(require_admin)):
    """Assign device to worker (admin only)"""
    body = await request.json()
    worker_id = body.get("worker_id")
    
    if not worker_id:
        raise HTTPException(status_code=400, detail="Wymagane worker_id")
    
    # Get device info before update
    device = await db.devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Nie znaleziono urządzenia")
    
    # Get worker info
    worker = await db.users.find_one({"user_id": worker_id}, {"_id": 0})
    worker_name = worker.get("name", "Nieznany") if worker else "Nieznany"
    
    result = await db.devices.update_one(
        {"device_id": device_id},
        {"$set": {"przypisany_do": worker_id, "status": "przypisany"}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono urządzenia")
    
    # Log assignment activity
    await log_activity(
        user_id=admin["user_id"],
        user_name=admin["name"],
        user_role="admin",
        action_type="device_assign",
        action_description=f"Przypisano urządzenie {device['nazwa']} ({device.get('numer_seryjny', 'brak SN')}) do {worker_name}",
        device_serial=device.get("numer_seryjny"),
        device_name=device["nazwa"],
        device_id=device_id,
        target_user_id=worker_id,
        target_user_name=worker_name
    )
    
    return {"message": "Urządzenie przypisane"}

@api_router.post("/devices/assign-multiple")
async def assign_multiple_devices(request: Request, admin: dict = Depends(require_admin)):
    """Assign multiple devices to a worker (admin only)"""
    body = await request.json()
    device_ids = body.get("device_ids", [])
    worker_id = body.get("worker_id")
    
    if not worker_id:
        raise HTTPException(status_code=400, detail="Wymagane worker_id")
    
    if not device_ids:
        raise HTTPException(status_code=400, detail="Wymagana lista device_ids")
    
    result = await db.devices.update_many(
        {"device_id": {"$in": device_ids}},
        {"$set": {"przypisany_do": worker_id, "status": "przypisany"}}
    )
    
    return {"message": f"Przypisano {result.modified_count} urządzeń"}

@api_router.post("/devices/{device_id}/restore")
async def restore_device(device_id: str, admin: dict = Depends(require_admin)):
    """Restore installed device back to available status for the original installer (admin only)"""
    # Check if device exists and is installed
    device = await db.devices.find_one({"device_id": device_id})
    if not device:
        raise HTTPException(status_code=404, detail="Nie znaleziono urządzenia")
    
    if device.get("status") != "zainstalowany":
        raise HTTPException(status_code=400, detail="Urządzenie nie jest zainstalowane")
    
    # Find the original installer from installations
    installation = await db.installations.find_one(
        {"device_id": device_id},
        sort=[("data_instalacji", -1)]
    )
    
    # Get the original installer - if not found, assign to admin
    original_installer = installation.get("user_id") if installation else admin["user_id"]
    
    # Restore device to available status and assign to original installer
    result = await db.devices.update_one(
        {"device_id": device_id},
        {"$set": {"status": "dostepny", "przypisany_do": original_installer}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=500, detail="Nie udało się przywrócić urządzenia")
    
    # Get installer name for response
    installer = await db.users.find_one({"user_id": original_installer})
    installer_name = installer.get("name", "Nieznany") if installer else "Nieznany"
    
    return {
        "message": f"Urządzenie zostało przywrócone do użytkownika: {installer_name}",
        "assigned_to": original_installer,
        "assigned_to_name": installer_name
    }

@api_router.post("/devices/{device_id}/transfer")
async def transfer_device(device_id: str, request: Request, admin: dict = Depends(require_admin)):
    """Transfer device from one worker to another (admin only)"""
    body = await request.json()
    new_worker_id = body.get("worker_id")
    
    if not new_worker_id:
        raise HTTPException(status_code=400, detail="Wymagane worker_id")
    
    # Check if device exists
    device = await db.devices.find_one({"device_id": device_id})
    if not device:
        raise HTTPException(status_code=404, detail="Nie znaleziono urządzenia")
    
    # Check if new worker exists
    new_worker = await db.users.find_one({"user_id": new_worker_id})
    if not new_worker:
        raise HTTPException(status_code=404, detail="Nie znaleziono pracownika")
    
    # Transfer device
    result = await db.devices.update_one(
        {"device_id": device_id},
        {"$set": {"przypisany_do": new_worker_id, "status": "przypisany"}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=500, detail="Nie udało się przenieść urządzenia")
    
    return {
        "message": f"Urządzenie przeniesione do: {new_worker.get('name', 'Nieznany')}",
        "new_worker_id": new_worker_id,
        "new_worker_name": new_worker.get("name")
    }

@api_router.get("/devices/scan/{code}")
async def scan_device(code: str, user: dict = Depends(require_user)):
    """Find device by barcode, QR code, or serial number (exact or partial match)"""
    # Clean the code - remove whitespace and special characters
    clean_code = code.strip().replace('\r', '').replace('\n', '')
    
    # First check if device is in pending returns
    in_returns = await db.device_returns.find_one({
        "device_serial": {"$regex": f"^{clean_code}$", "$options": "i"},
        "returned_to_warehouse": {"$ne": True}
    })
    if in_returns:
        raise HTTPException(
            status_code=400, 
            detail=f"Urządzenie {clean_code} jest już w panelu Zwrot urządzeń"
        )
    
    # First try exact match
    device = await db.devices.find_one(
        {"$or": [
            {"kod_kreskowy": clean_code}, 
            {"kod_qr": clean_code}, 
            {"numer_seryjny": clean_code}
        ]},
        {"_id": 0}
    )
    
    if not device:
        # Try case-insensitive match
        device = await db.devices.find_one(
            {"$or": [
                {"kod_kreskowy": {"$regex": f"^{clean_code}$", "$options": "i"}}, 
                {"kod_qr": {"$regex": f"^{clean_code}$", "$options": "i"}}, 
                {"numer_seryjny": {"$regex": f"^{clean_code}$", "$options": "i"}}
            ]},
            {"_id": 0}
        )
    
    if not device:
        # Try partial match on serial number (contains)
        device = await db.devices.find_one(
            {"numer_seryjny": {"$regex": clean_code, "$options": "i"}},
            {"_id": 0}
        )
    
    if not device:
        # Try if code contains the serial number
        all_devices = await db.devices.find({}, {"_id": 0}).to_list(1000)
        for dev in all_devices:
            if dev.get("numer_seryjny") and dev["numer_seryjny"].upper() in clean_code.upper():
                device = dev
                break
            if dev.get("kod_kreskowy") and dev["kod_kreskowy"].upper() in clean_code.upper():
                device = dev
                break
            if dev.get("kod_qr") and dev["kod_qr"].upper() in clean_code.upper():
                device = dev
                break
    
    if not device:
        raise HTTPException(status_code=404, detail="Nie znaleziono urządzenia")
    
    # Check if device is already installed or damaged - workers can't scan these again
    if user.get("role") != "admin":
        if device.get("status") == "zainstalowany":
            raise HTTPException(
                status_code=400, 
                detail="To urządzenie jest już zainstalowane u klienta"
            )
        if device.get("status") == "uszkodzony":
            raise HTTPException(
                status_code=400, 
                detail="To urządzenie jest oznaczone jako uszkodzone"
            )
        if device.get("status") == "zwrocony":
            raise HTTPException(
                status_code=400, 
                detail="To urządzenie zostało zwrócone do magazynu"
            )
    
    return device

# ==================== INSTALLATIONS ====================

@api_router.post("/installations")
async def create_installation(request: Request, user: dict = Depends(require_user)):
    """Record device installation - moves device to admin account as installed"""
    body = await request.json()
    
    device_id = body.get("device_id")
    adres_klienta = body.get("adres_klienta") or body.get("adres")
    
    if not device_id:
        raise HTTPException(status_code=400, detail="Wymagane device_id")
    
    if not adres_klienta or not adres_klienta.strip():
        raise HTTPException(status_code=400, detail="Wymagany adres klienta")
    
    device = await db.devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Nie znaleziono urządzenia")
    
    # Check if device is assigned to this user
    if device.get("przypisany_do") != user["user_id"] and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="To urządzenie nie jest przypisane do Ciebie")
    
    # Get admin user
    admin_user = await db.users.find_one({"role": "admin"}, {"_id": 0})
    if not admin_user:
        raise HTTPException(status_code=500, detail="Brak administratora w systemie")
    
    installation = {
        "installation_id": f"inst_{uuid.uuid4().hex[:12]}",
        "device_id": device_id,
        "user_id": user["user_id"],
        "installer_name": user["name"],
        "nazwa_urzadzenia": device["nazwa"],
        "numer_seryjny": device.get("numer_seryjny", ""),
        "kod_kreskowy": device.get("kod_kreskowy", ""),
        "data_instalacji": datetime.now(timezone.utc),
        "adres_klienta": adres_klienta.strip(),
        "latitude": body.get("latitude"),
        "longitude": body.get("longitude"),
        "rodzaj_zlecenia": body.get("rodzaj_zlecenia", "instalacja")
    }
    
    await db.installations.insert_one(installation)
    
    # Move device to admin account as installed
    await db.devices.update_one(
        {"device_id": device_id},
        {"$set": {
            "status": "zainstalowany",
            "przypisany_do": admin_user["user_id"],
            "zainstalowany_przez": user["user_id"],
            "installer_name": user["name"],
            "adres_instalacji": adres_klienta.strip(),
            "data_instalacji": datetime.now(timezone.utc)
        }}
    )
    
    # Log installation activity
    await log_activity(
        user_id=user["user_id"],
        user_name=user["name"],
        user_role=user.get("role", "pracownik"),
        action_type="device_install",
        action_description=f"Zainstalowano urządzenie {device['nazwa']} ({device.get('numer_seryjny', 'brak SN')})",
        device_serial=device.get("numer_seryjny"),
        device_name=device["nazwa"],
        device_id=device_id,
        details={
            "adres_klienta": adres_klienta.strip(),
            "rodzaj_zlecenia": body.get("rodzaj_zlecenia", "instalacja"),
            "latitude": body.get("latitude"),
            "longitude": body.get("longitude")
        }
    )
    
    installation.pop("_id", None)
    return installation

@api_router.get("/installations")
async def get_installations(
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    rodzaj_zlecenia: Optional[str] = None,
    current_user: dict = Depends(require_user)
):
    """Get installations with filters"""
    query = {}
    
    if user_id:
        query["user_id"] = user_id
    elif current_user.get("role") != "admin":
        query["user_id"] = current_user["user_id"]
    
    if rodzaj_zlecenia:
        query["rodzaj_zlecenia"] = rodzaj_zlecenia
    
    if date_from:
        query["data_instalacji"] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        if "data_instalacji" in query:
            query["data_instalacji"]["$lte"] = datetime.fromisoformat(date_to)
        else:
            query["data_instalacji"] = {"$lte": datetime.fromisoformat(date_to)}
    
    installations = await db.installations.find(query, {"_id": 0}).to_list(1000)
    return installations

@api_router.get("/installations/stats")
async def get_installation_stats(user: dict = Depends(require_user)):
    """Get installation statistics"""
    pipeline = [
        {"$group": {
            "_id": "$rodzaj_zlecenia",
            "count": {"$sum": 1}
        }}
    ]
    
    stats_by_type = await db.installations.aggregate(pipeline).to_list(100)
    
    pipeline_users = [
        {"$group": {
            "_id": "$user_id",
            "count": {"$sum": 1}
        }}
    ]
    stats_by_user = await db.installations.aggregate(pipeline_users).to_list(100)
    
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    pipeline_daily = [
        {"$match": {"data_instalacji": {"$gte": week_ago}}},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$data_instalacji"}},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]
    stats_daily = await db.installations.aggregate(pipeline_daily).to_list(100)
    
    total = await db.installations.count_documents({})
    
    return {
        "total": total,
        "by_type": {item["_id"]: item["count"] for item in stats_by_type if item["_id"]},
        "by_user": {item["_id"]: item["count"] for item in stats_by_user if item["_id"]},
        "daily": stats_daily
    }

# ==================== MESSAGES / CHAT ====================

@api_router.post("/messages")
async def send_message(request: Request, user: dict = Depends(require_user)):
    """Send a message"""
    body = await request.json()
    
    message = {
        "message_id": f"msg_{uuid.uuid4().hex[:12]}",
        "sender_id": user["user_id"],
        "sender_name": user["name"],
        "content": body.get("content"),
        "attachment": body.get("attachment"),
        "attachment_type": body.get("attachment_type"),
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.messages.insert_one(message)
    message.pop("_id", None)
    return message

@api_router.get("/messages")
async def get_messages(
    limit: int = 50,
    before: Optional[str] = None,
    user: dict = Depends(require_user)
):
    """Get messages"""
    query = {}
    if before:
        query["created_at"] = {"$lt": datetime.fromisoformat(before)}
    
    messages = await db.messages.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return list(reversed(messages))

# ==================== TASKS / PLANNER ====================

@api_router.post("/tasks")
async def create_task(request: Request, admin: dict = Depends(require_admin)):
    """Create a task (admin only)"""
    body = await request.json()
    
    task = {
        "task_id": f"task_{uuid.uuid4().hex[:12]}",
        "title": body.get("title"),
        "description": body.get("description"),
        "assigned_to": body.get("assigned_to"),
        "assigned_by": admin["user_id"],
        "due_date": datetime.fromisoformat(body.get("due_date")) if body.get("due_date") else datetime.now(timezone.utc),
        "status": "oczekujace",
        "priority": body.get("priority", "normalne"),
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.tasks.insert_one(task)
    task.pop("_id", None)
    return task

@api_router.get("/tasks")
async def get_tasks(
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    user: dict = Depends(require_user)
):
    """Get tasks"""
    query = {}
    
    if status:
        query["status"] = status
    
    if user.get("role") != "admin":
        query["assigned_to"] = user["user_id"]
    elif assigned_to:
        query["assigned_to"] = assigned_to
    
    tasks = await db.tasks.find(query, {"_id": 0}).sort("due_date", 1).to_list(1000)
    return tasks

@api_router.put("/tasks/{task_id}")
async def update_task(task_id: str, request: Request, user: dict = Depends(require_user)):
    """Update task status"""
    body = await request.json()
    
    update_data = {}
    if "status" in body:
        update_data["status"] = body["status"]
    if "title" in body and user.get("role") == "admin":
        update_data["title"] = body["title"]
    if "description" in body and user.get("role") == "admin":
        update_data["description"] = body["description"]
    if "due_date" in body and user.get("role") == "admin":
        update_data["due_date"] = datetime.fromisoformat(body["due_date"])
    if "priority" in body and user.get("role") == "admin":
        update_data["priority"] = body["priority"]
    if "completion_photos" in body:
        update_data["completion_photos"] = body["completion_photos"]
        update_data["completed_at"] = datetime.now(timezone.utc)
        update_data["completed_by"] = user["user_id"]
    
    result = await db.tasks.update_one(
        {"task_id": task_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono zadania")
    
    return {"message": "Zadanie zaktualizowane"}

@api_router.get("/tasks/{task_id}")
async def get_task(task_id: str, user: dict = Depends(require_user)):
    """Get single task with full details"""
    task = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Nie znaleziono zadania")
    
    # Check permissions - workers can only see their own tasks
    if user.get("role") != "admin" and task.get("assigned_to") != user["user_id"]:
        raise HTTPException(status_code=403, detail="Brak uprawnień")
    
    return task

@api_router.get("/tasks/reminders/check")
async def check_task_reminders(user: dict = Depends(require_user)):
    """Check for tasks approaching deadline (within 2 hours)"""
    now = datetime.now(timezone.utc)
    two_hours_later = now + timedelta(hours=2)
    
    query = {
        "status": {"$ne": "zakonczone"},
        "due_date": {"$lte": two_hours_later, "$gt": now}
    }
    
    # Workers only see their own reminders
    if user.get("role") != "admin":
        query["assigned_to"] = user["user_id"]
    
    tasks = await db.tasks.find(query, {"_id": 0}).to_list(100)
    
    reminders = []
    for task in tasks:
        due_date = task.get("due_date")
        if isinstance(due_date, datetime):
            # Ensure due_date is timezone aware
            if due_date.tzinfo is None:
                due_date = due_date.replace(tzinfo=timezone.utc)
            time_left = due_date - now
            minutes_left = int(time_left.total_seconds() / 60)
            
            reminders.append({
                "task_id": task["task_id"],
                "title": task["title"],
                "due_date": due_date.isoformat(),
                "minutes_left": minutes_left,
                "assigned_to": task.get("assigned_to"),
                "priority": task.get("priority", "normalne")
            })
    
    return {"reminders": reminders, "count": len(reminders)}

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, admin: dict = Depends(require_admin)):
    """Delete task (admin only)"""
    result = await db.tasks.delete_one({"task_id": task_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono zadania")
    return {"message": "Zadanie usunięte"}

# ==================== DAILY REPORT ====================

@api_router.get("/report/daily")
async def get_daily_report(user: dict = Depends(require_user)):
    """Get daily installations report"""
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)
    
    installations = await db.installations.find(
        {"data_instalacji": {"$gte": today, "$lt": tomorrow}},
        {"_id": 0}
    ).to_list(1000)
    
    by_user = {}
    for inst in installations:
        uid = inst["user_id"]
        if uid not in by_user:
            by_user[uid] = []
        by_user[uid].append(inst)
    
    report = []
    for uid, insts in by_user.items():
        user_doc = await db.users.find_one({"user_id": uid}, {"_id": 0, "password_hash": 0})
        report.append({
            "user_id": uid,
            "user_name": user_doc["name"] if user_doc else "Nieznany",
            "count": len(insts),
            "installations": insts
        })
    
    return {
        "date": today.isoformat(),
        "total": len(installations),
        "by_user": report
    }

# ==================== BACKUP FUNCTIONS ====================

async def create_backup_data() -> dict:
    """Create backup of all database collections"""
    backup = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "version": "1.0",
        "data": {}
    }
    
    # Backup users (without passwords for security, but with hashed passwords for restore)
    users = await db.users.find({}, {"_id": 0}).to_list(1000)
    backup["data"]["users"] = users
    
    # Backup devices
    devices = await db.devices.find({}, {"_id": 0}).to_list(10000)
    backup["data"]["devices"] = devices
    
    # Backup installations
    installations = await db.installations.find({}, {"_id": 0}).to_list(10000)
    # Convert datetime objects to strings
    for inst in installations:
        if "data_instalacji" in inst and isinstance(inst["data_instalacji"], datetime):
            inst["data_instalacji"] = inst["data_instalacji"].isoformat()
    backup["data"]["installations"] = installations
    
    # Backup tasks
    tasks = await db.tasks.find({}, {"_id": 0}).to_list(10000)
    for task in tasks:
        if "due_date" in task and isinstance(task["due_date"], datetime):
            task["due_date"] = task["due_date"].isoformat()
        if "created_at" in task and isinstance(task["created_at"], datetime):
            task["created_at"] = task["created_at"].isoformat()
    backup["data"]["tasks"] = tasks
    
    # Backup messages (without attachments for size)
    messages = await db.messages.find({}, {"_id": 0, "attachment": 0}).to_list(10000)
    for msg in messages:
        if "created_at" in msg and isinstance(msg["created_at"], datetime):
            msg["created_at"] = msg["created_at"].isoformat()
    backup["data"]["messages"] = messages
    
    return backup

def send_backup_email(backup_data: bytes, filename: str, settings: dict) -> bool:
    """Send backup file via email"""
    try:
        msg = MIMEMultipart()
        msg['From'] = settings['smtp_user']
        msg['To'] = settings['email_recipient']
        msg['Subject'] = f"Kopia zapasowa Magazyn ITS - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        body = f"""Automatyczna kopia zapasowa bazy danych Magazyn ITS Kielce.

Data utworzenia: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Rozmiar: {len(backup_data) / 1024:.2f} KB

Ta wiadomość została wygenerowana automatycznie."""
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # Attach backup file
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(backup_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
        
        # Send email
        if settings.get('smtp_use_tls', True):
            server = smtplib.SMTP(settings['smtp_host'], settings['smtp_port'])
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(settings['smtp_host'], settings['smtp_port'])
        
        server.login(settings['smtp_user'], settings['smtp_password'])
        server.send_message(msg)
        server.quit()
        
        return True
    except Exception as e:
        logger.error(f"Failed to send backup email: {e}")
        return False

def send_backup_ftp(backup_data: bytes, filename: str, settings: dict) -> bool:
    """Upload backup file to FTP server"""
    try:
        ftp = ftplib.FTP()
        ftp.connect(settings['ftp_host'], settings.get('ftp_port', 21))
        ftp.login(settings['ftp_user'], settings['ftp_password'])
        
        # Navigate to backup directory
        ftp_path = settings.get('ftp_path', '/backups/')
        try:
            ftp.cwd(ftp_path)
        except:
            # Try to create directory if it doesn't exist
            ftp.mkd(ftp_path)
            ftp.cwd(ftp_path)
        
        # Upload file
        ftp.storbinary(f'STOR {filename}', BytesIO(backup_data))
        ftp.quit()
        
        return True
    except Exception as e:
        logger.error(f"Failed to upload backup to FTP: {e}")
        return False

# ==================== BACKUP ENDPOINTS ====================

@api_router.get("/backup/settings")
async def get_backup_settings(request: Request):
    """Get backup settings (admin only)"""
    user = await get_current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Brak uprawnień")
    
    settings = await db.backup_settings.find_one({}, {"_id": 0})
    if not settings:
        # Return default settings
        settings = {
            "smtp_host": "",
            "smtp_port": 587,
            "smtp_user": "",
            "smtp_password": "",
            "smtp_use_tls": True,
            "email_recipient": "",
            "email_enabled": False,
            "ftp_host": "",
            "ftp_port": 21,
            "ftp_user": "",
            "ftp_password": "",
            "ftp_path": "/backups/",
            "ftp_enabled": False,
            "schedule_enabled": False,
            "schedule_time": "02:00"
        }
    
    # Hide passwords in response
    if settings.get("smtp_password"):
        settings["smtp_password"] = "********"
    if settings.get("ftp_password"):
        settings["ftp_password"] = "********"
    
    return settings

@api_router.post("/backup/settings")
async def update_backup_settings(request: Request):
    """Update backup settings (admin only)"""
    user = await get_current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Brak uprawnień")
    
    data = await request.json()
    
    # Get existing settings to preserve passwords if not changed
    existing = await db.backup_settings.find_one({})
    
    settings = {
        "smtp_host": data.get("smtp_host", ""),
        "smtp_port": data.get("smtp_port", 587),
        "smtp_user": data.get("smtp_user", ""),
        "smtp_use_tls": data.get("smtp_use_tls", True),
        "email_recipient": data.get("email_recipient", ""),
        "email_enabled": data.get("email_enabled", False),
        "ftp_host": data.get("ftp_host", ""),
        "ftp_port": data.get("ftp_port", 21),
        "ftp_user": data.get("ftp_user", ""),
        "ftp_path": data.get("ftp_path", "/backups/"),
        "ftp_enabled": data.get("ftp_enabled", False),
        "schedule_enabled": data.get("schedule_enabled", False),
        "schedule_time": data.get("schedule_time", "02:00"),
        "updated_at": datetime.now(timezone.utc)
    }
    
    # Handle password fields - only update if not placeholder
    if data.get("smtp_password") and data.get("smtp_password") != "********":
        settings["smtp_password"] = data["smtp_password"]
    elif existing and existing.get("smtp_password"):
        settings["smtp_password"] = existing["smtp_password"]
    else:
        settings["smtp_password"] = ""
    
    if data.get("ftp_password") and data.get("ftp_password") != "********":
        settings["ftp_password"] = data["ftp_password"]
    elif existing and existing.get("ftp_password"):
        settings["ftp_password"] = existing["ftp_password"]
    else:
        settings["ftp_password"] = ""
    
    await db.backup_settings.update_one({}, {"$set": settings}, upsert=True)
    
    return {"status": "ok", "message": "Ustawienia zostały zapisane"}

@api_router.post("/backup/create")
async def create_backup(request: Request):
    """Create a new backup (admin only)"""
    user = await get_current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Brak uprawnień")
    
    data = await request.json()
    send_email = data.get("send_email", False)
    send_ftp = data.get("send_ftp", False)
    
    try:
        # Create backup data
        backup = await create_backup_data()
        backup_json = json.dumps(backup, ensure_ascii=False, indent=2)
        backup_bytes = backup_json.encode('utf-8')
        
        filename = f"magazyn_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        # Log backup
        log = {
            "backup_id": f"backup_{uuid.uuid4().hex[:12]}",
            "created_at": datetime.now(timezone.utc),
            "size_bytes": len(backup_bytes),
            "status": "success",
            "sent_email": False,
            "sent_ftp": False,
            "error_message": None
        }
        
        errors = []
        
        # Send via email if requested
        if send_email:
            settings = await db.backup_settings.find_one({})
            if settings and settings.get("email_enabled") and settings.get("smtp_host"):
                if send_backup_email(backup_bytes, filename, settings):
                    log["sent_email"] = True
                else:
                    errors.append("Email: nie udało się wysłać")
            else:
                errors.append("Email: nie skonfigurowano")
        
        # Send via FTP if requested
        if send_ftp:
            settings = await db.backup_settings.find_one({})
            if settings and settings.get("ftp_enabled") and settings.get("ftp_host"):
                if send_backup_ftp(backup_bytes, filename, settings):
                    log["sent_ftp"] = True
                else:
                    errors.append("FTP: nie udało się wysłać")
            else:
                errors.append("FTP: nie skonfigurowano")
        
        if errors:
            log["error_message"] = "; ".join(errors)
        
        await db.backup_logs.insert_one(log)
        
        return {
            "status": "ok",
            "backup_id": log["backup_id"],
            "size_bytes": log["size_bytes"],
            "sent_email": log["sent_email"],
            "sent_ftp": log["sent_ftp"],
            "errors": errors
        }
        
    except Exception as e:
        logger.error(f"Backup failed: {e}")
        raise HTTPException(status_code=500, detail=f"Błąd tworzenia kopii zapasowej: {str(e)}")

@api_router.get("/backup/download")
async def download_backup(request: Request):
    """Download backup file (admin only)"""
    user = await get_current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Brak uprawnień")
    
    try:
        backup = await create_backup_data()
        backup_json = json.dumps(backup, ensure_ascii=False, indent=2)
        backup_bytes = backup_json.encode('utf-8')
        
        filename = f"magazyn_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        # Log the download
        log = {
            "backup_id": f"backup_{uuid.uuid4().hex[:12]}",
            "created_at": datetime.now(timezone.utc),
            "size_bytes": len(backup_bytes),
            "status": "success",
            "sent_email": False,
            "sent_ftp": False,
            "downloaded": True
        }
        await db.backup_logs.insert_one(log)
        
        return StreamingResponse(
            BytesIO(backup_bytes),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    except Exception as e:
        logger.error(f"Backup download failed: {e}")
        raise HTTPException(status_code=500, detail=f"Błąd pobierania kopii zapasowej: {str(e)}")

@api_router.get("/backup/logs")
async def get_backup_logs(request: Request):
    """Get backup history (admin only)"""
    user = await get_current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Brak uprawnień")
    
    logs = await db.backup_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return logs

@api_router.post("/backup/test-email")
async def test_email_backup(request: Request):
    """Test email configuration (admin only)"""
    user = await get_current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Brak uprawnień")
    
    settings = await db.backup_settings.find_one({})
    if not settings or not settings.get("smtp_host"):
        raise HTTPException(status_code=400, detail="Email nie jest skonfigurowany")
    
    try:
        # Create a small test backup
        test_data = {
            "test": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "message": "To jest testowa kopia zapasowa"
        }
        test_bytes = json.dumps(test_data, ensure_ascii=False).encode('utf-8')
        filename = f"test_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        if send_backup_email(test_bytes, filename, settings):
            return {"status": "ok", "message": "Email testowy został wysłany"}
        else:
            raise HTTPException(status_code=500, detail="Nie udało się wysłać emaila testowego")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Błąd: {str(e)}")

@api_router.post("/backup/test-ftp")
async def test_ftp_backup(request: Request):
    """Test FTP configuration (admin only)"""
    user = await get_current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Brak uprawnień")
    
    settings = await db.backup_settings.find_one({})
    if not settings or not settings.get("ftp_host"):
        raise HTTPException(status_code=400, detail="FTP nie jest skonfigurowany")
    
    try:
        # Create a small test file
        test_data = {
            "test": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "message": "To jest testowa kopia zapasowa"
        }
        test_bytes = json.dumps(test_data, ensure_ascii=False).encode('utf-8')
        filename = f"test_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        if send_backup_ftp(test_bytes, filename, settings):
            return {"status": "ok", "message": "Plik testowy został wysłany na FTP"}
        else:
            raise HTTPException(status_code=500, detail="Nie udało się wysłać pliku na FTP")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Błąd: {str(e)}")

# ==================== DEVICE RETURNS ====================

class DeviceReturn(BaseModel):
    return_id: str = Field(default_factory=lambda: f"ret_{uuid.uuid4().hex[:12]}")
    device_serial: str
    device_type: str  # ONT, CPE, STB
    device_status: str  # z awarii, nowy/uszkodzony
    scanned_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    scanned_by: str
    scanned_by_name: str

@api_router.post("/returns")
async def add_device_return(request: Request, admin: dict = Depends(require_admin)):
    """Add a device to returns (admin only)"""
    body = await request.json()
    
    device_serial = body.get("device_serial")
    device_type = body.get("device_type")
    device_status = body.get("device_status")
    
    if not device_serial:
        raise HTTPException(status_code=400, detail="Numer seryjny jest wymagany")
    if not device_type:
        raise HTTPException(status_code=400, detail="Rodzaj urządzenia jest wymagany")
    if not device_status:
        raise HTTPException(status_code=400, detail="Stan urządzenia jest wymagany")
    
    # Check for duplicates (only in pending returns)
    existing = await db.device_returns.find_one({"device_serial": device_serial, "returned_to_warehouse": {"$ne": True}})
    if existing:
        raise HTTPException(status_code=400, detail="Ten numer seryjny już jest w zwrotach")
    
    return_entry = {
        "return_id": f"ret_{uuid.uuid4().hex[:12]}",
        "device_serial": device_serial,
        "device_type": device_type,
        "device_status": device_status,
        "scanned_at": datetime.now(timezone.utc),
        "scanned_by": admin["user_id"],
        "scanned_by_name": admin["name"]
    }
    
    await db.device_returns.insert_one(return_entry)
    return_entry.pop("_id", None)
    
    return return_entry

@api_router.get("/returns")
async def get_device_returns(admin: dict = Depends(require_admin)):
    """Get all device returns (admin only)"""
    returns = await db.device_returns.find({}, {"_id": 0}).sort("scanned_at", -1).to_list(10000)
    return returns

@api_router.delete("/returns/{return_id}")
async def delete_device_return(return_id: str, admin: dict = Depends(require_admin)):
    """Delete a device return entry (admin only)"""
    result = await db.device_returns.delete_one({"return_id": return_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono wpisu")
    return {"message": "Wpis usunięty"}

@api_router.post("/returns/bulk")
async def add_bulk_returns(request: Request, admin: dict = Depends(require_admin)):
    """Add multiple devices to returns from damaged devices (admin only)"""
    body = await request.json()
    device_serials = body.get("device_serials", [])
    device_type = body.get("device_type", "")  # Can be empty, admin fills later
    device_status = body.get("device_status", "nowy/uszkodzony")
    
    if not device_serials:
        raise HTTPException(status_code=400, detail="Brak urządzeń do dodania")
    
    added = 0
    skipped = 0
    for serial in device_serials:
        # Check for duplicates
        existing = await db.device_returns.find_one({"device_serial": serial, "returned_to_warehouse": {"$ne": True}})
        if existing:
            skipped += 1
            continue
            
        return_entry = {
            "return_id": f"ret_{uuid.uuid4().hex[:12]}",
            "device_serial": serial,
            "device_type": device_type,
            "device_status": device_status,
            "scanned_at": datetime.now(timezone.utc),
            "scanned_by": admin["user_id"],
            "scanned_by_name": admin["name"]
        }
        await db.device_returns.insert_one(return_entry)
        
        # Remove device from employee's account (change status to 'zwrocony' or delete assignment)
        await db.devices.update_one(
            {"numer_seryjny": serial},
            {"$set": {
                "status": "zwrocony",
                "przypisany_do": None,
                "returned_at": datetime.now(timezone.utc),
                "returned_by": admin["user_id"]
            }}
        )
        added += 1
    
    message = f"Dodano {added} urządzeń do zwrotów"
    if skipped > 0:
        message += f" (pominięto {skipped} duplikatów)"
    
    return {"message": message, "added": added, "skipped": skipped}

@api_router.put("/returns/{return_id}")
async def update_device_return(return_id: str, request: Request, admin: dict = Depends(require_admin)):
    """Update a device return entry (admin only)"""
    body = await request.json()
    
    update_data = {}
    if "device_type" in body:
        update_data["device_type"] = body["device_type"]
    if "device_status" in body:
        update_data["device_status"] = body["device_status"]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Brak danych do aktualizacji")
    
    result = await db.device_returns.update_one(
        {"return_id": return_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono wpisu")
    
    return {"message": "Wpis zaktualizowany"}

@api_router.get("/returns/export")
async def export_returns_excel(request: Request, token: str = None):
    """Export device returns to Excel (admin only)"""
    # Try to get admin from header first, then from query param for mobile
    admin = None
    
    # Try header authentication
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        session_token = auth_header.replace("Bearer ", "")
        session = await db.user_sessions.find_one({"session_token": session_token})
        if session:
            admin = await db.users.find_one({"user_id": session["user_id"]})
    
    # Try query param authentication (for mobile)
    if not admin and token:
        session = await db.user_sessions.find_one({"session_token": token})
        if session:
            admin = await db.users.find_one({"user_id": session["user_id"]})
    
    if not admin or admin.get("role") != "admin":
        raise HTTPException(status_code=401, detail="Brak uprawnień")
    
    returns = await db.device_returns.find(
        {"returned_to_warehouse": {"$ne": True}},  # Only pending returns
        {"_id": 0}
    ).sort("scanned_at", -1).to_list(10000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zwroty urządzeń"
    
    # Headers
    headers = ["Numer seryjny", "Rodzaj", "Stan", "Data skanowania"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row_num, ret in enumerate(returns, 2):
        ws.cell(row=row_num, column=1, value=ret.get("device_serial", ""))
        ws.cell(row=row_num, column=2, value=ret.get("device_type", ""))
        ws.cell(row=row_num, column=3, value=ret.get("device_status", ""))
        scanned_at = ret.get("scanned_at")
        if isinstance(scanned_at, datetime):
            ws.cell(row=row_num, column=4, value=scanned_at.strftime("%d-%m-%Y"))
        else:
            ws.cell(row=row_num, column=4, value=str(scanned_at) if scanned_at else "")
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"zwroty_urzadzen_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/returns/mark-returned")
async def mark_returns_as_returned(admin: dict = Depends(require_admin)):
    """Mark all pending returns as returned to warehouse (admin only)"""
    result = await db.device_returns.update_many(
        {"returned_to_warehouse": {"$ne": True}},
        {"$set": {
            "returned_to_warehouse": True,
            "returned_at": datetime.now(timezone.utc)
        }}
    )
    
    return {
        "message": f"Oznaczono {result.modified_count} urządzeń jako zwrócone do magazynu",
        "count": result.modified_count
    }

# ==================== DEVICE STATUS UPDATE - DAMAGED ====================

@api_router.post("/devices/{device_id}/mark-damaged")
async def mark_device_damaged(device_id: str, request: Request, user: dict = Depends(require_user)):
    """Mark a device as damaged"""
    body = await request.json()
    
    device = await db.devices.find_one({"device_id": device_id})
    if not device:
        raise HTTPException(status_code=404, detail="Nie znaleziono urządzenia")
    
    # Check if user has access to this device
    if user.get("role") != "admin" and device.get("przypisany_do") != user["user_id"]:
        raise HTTPException(status_code=403, detail="Brak uprawnień do tego urządzenia")
    
    # Update device status to damaged
    await db.devices.update_one(
        {"device_id": device_id},
        {"$set": {
            "status": "uszkodzony",
            "damaged_at": datetime.now(timezone.utc),
            "damaged_by": user["user_id"],
            "damaged_by_name": user["name"]
        }}
    )
    
    return {"message": "Urządzenie oznaczone jako uszkodzone"}

@api_router.post("/devices/add-single")
async def add_single_device(request: Request, admin: dict = Depends(require_admin)):
    """Add a single device manually via barcode scan (admin only)"""
    body = await request.json()
    
    nazwa = body.get("nazwa", "")
    numer_seryjny = body.get("numer_seryjny")
    kod_kreskowy = body.get("kod_kreskowy")
    
    if not numer_seryjny and not kod_kreskowy:
        raise HTTPException(status_code=400, detail="Wymagany numer seryjny lub kod kreskowy")
    
    # Use barcode as serial if serial not provided
    if not numer_seryjny:
        numer_seryjny = kod_kreskowy
    
    # Check if device already exists
    existing = await db.devices.find_one({"numer_seryjny": numer_seryjny})
    if existing:
        raise HTTPException(status_code=400, detail="Urządzenie o tym numerze seryjnym już istnieje")
    
    device = {
        "device_id": f"dev_{uuid.uuid4().hex[:12]}",
        "nazwa": nazwa,
        "numer_seryjny": numer_seryjny,
        "kod_kreskowy": kod_kreskowy or numer_seryjny,
        "kod_qr": body.get("kod_qr"),
        "przypisany_do": None,
        "status": "dostepny",
        "created_at": datetime.now(timezone.utc),
        "added_by": admin["user_id"],
        "added_manually": True
    }
    
    await db.devices.insert_one(device)
    device.pop("_id", None)
    
    # Log device addition activity
    await log_activity(
        user_id=admin["user_id"],
        user_name=admin["name"],
        user_role="admin",
        action_type="device_add",
        action_description=f"Dodano nowe urządzenie {nazwa} ({numer_seryjny}) do magazynu",
        device_serial=numer_seryjny,
        device_name=nazwa,
        device_id=device["device_id"]
    )
    
    return device

# ==================== ACTIVITY LOGS ENDPOINTS ====================

@api_router.get("/activity-logs/user/{user_id}")
async def get_user_activity_logs(
    user_id: str,
    limit: int = 100,
    admin: dict = Depends(require_admin)
):
    """Get activity logs for a specific user (admin only)"""
    logs = await db.activity_logs.find(
        {"user_id": user_id},
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit).to_list(limit)
    
    # Format timestamps for JSON serialization
    for log in logs:
        if "timestamp" in log and isinstance(log["timestamp"], datetime):
            log["timestamp"] = log["timestamp"].isoformat()
    
    return logs

@api_router.get("/activity-logs/device/{device_serial}")
async def get_device_history(
    device_serial: str,
    limit: int = 100,
    admin: dict = Depends(require_admin)
):
    """Get activity history for a specific device serial number (admin only)"""
    logs = await db.activity_logs.find(
        {"device_serial": device_serial},
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit).to_list(limit)
    
    # Format timestamps for JSON serialization
    for log in logs:
        if "timestamp" in log and isinstance(log["timestamp"], datetime):
            log["timestamp"] = log["timestamp"].isoformat()
    
    return logs

@api_router.get("/activity-logs/recent")
async def get_recent_activity_logs(
    limit: int = 100,
    user_id: Optional[str] = None,
    action_type: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Get recent activity logs with optional filters (admin only)"""
    query = {}
    if user_id:
        query["user_id"] = user_id
    if action_type:
        query["action_type"] = action_type
    
    logs = await db.activity_logs.find(
        query,
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit).to_list(limit)
    
    # Format timestamps for JSON serialization
    for log in logs:
        if "timestamp" in log and isinstance(log["timestamp"], datetime):
            log["timestamp"] = log["timestamp"].isoformat()
    
    return logs

# ==================== HEALTH CHECK ====================

@api_router.get("/")
async def root():
    return {"message": "Magazyn ITS Kielce API", "status": "ok"}

@api_router.get("/health")
async def health():
    return {"status": "ok"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
